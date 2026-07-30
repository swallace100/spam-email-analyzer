#!/usr/bin/env python3
"""Regenerate output/master_report.xlsx from the running master CSV.
Safe to run repeatedly -- it's a full rebuild from the CSV, not an append,
so the CSV is always the source of truth."""

import csv
import email.utils
import ipaddress
import os
from collections import Counter, defaultdict
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
MASTER_CSV = os.path.join(BASE, "data", "master_iocs.csv")
MASTER_URLS_CSV = os.path.join(BASE, "data", "master_urls.csv")
ENRICHMENT_CSV = os.path.join(BASE, "data", "domain_enrichment.csv")
WALLET_ENRICHMENT_CSV = os.path.join(BASE, "data", "wallet_enrichment.csv")
IP_ENRICHMENT_CSV = os.path.join(BASE, "data", "ip_enrichment.csv")
OUT_XLSX = os.path.join(BASE, "output", "master_report.xlsx")


def load_wallet_enrichment():
    """On-chain activity per wallet, built by lookup_wallets.py."""
    if not os.path.exists(WALLET_ENRICHMENT_CSV):
        return {}
    with open(WALLET_ENRICHMENT_CSV, encoding="utf-8") as f:
        return {r["address"]: r for r in csv.DictReader(f)}


def load_ip_enrichment():
    """ASN/provider info per IP, built by lookup_ips.py."""
    if not os.path.exists(IP_ENRICHMENT_CSV):
        return {}
    with open(IP_ENRICHMENT_CSV, encoding="utf-8") as f:
        return {r["ip"]: r for r in csv.DictReader(f)}


def origin_ip_of(row):
    """Best guess at the true origin IP of a message: the LAST public IP in
    originating_ips. extract_ips preserves Received-header order (top =
    closest to your inbox, bottom = closest to the sender), so the last
    entry is the deepest hop the chain records. Rows ingested before IP
    validation may contain junk, so entries are validated here too."""
    candidates = [ip.strip() for ip in (row.get("originating_ips") or "").split(";") if ip.strip()]
    for ip in reversed(candidates):
        try:
            if ipaddress.ip_address(ip).is_global:
                return ip
        except ValueError:
            continue
    return ""


def month_of(row):
    """'YYYY-MM' for a row, preferring the normalized date_iso column and
    falling back to parsing the raw Date header (for rows ingested before
    date_iso existed)."""
    d = (row.get("date_iso") or "").strip()
    if len(d) >= 7:
        return d[:7]
    try:
        dt = email.utils.parsedate_to_datetime(row.get("date", ""))
        return dt.strftime("%Y-%m") if dt else ""
    except Exception:
        return ""


def load_enrichment():
    if not os.path.exists(ENRICHMENT_CSV):
        return {}
    with open(ENRICHMENT_CSV, encoding="utf-8") as f:
        return {r["domain"]: r for r in csv.DictReader(f)}


# --- Template clustering (groups messages by body similarity) ----------
# ioc_lib stores a SimHash fingerprint of each body's template (URLs/
# emails/phones/wallets stripped out) rather than the body itself. Two
# fingerprints with a small Hamming distance likely came from the same
# scam template even if sender and contact info differ. O(n^2) pairwise
# comparison is fine at the personal-mailbox scale this tool targets
# (thousands of rows); if that ever changes, band the fingerprints instead
# of comparing every pair.
TEMPLATE_HAMMING_THRESHOLD = 3


def _hamming(a, b):
    return bin(a ^ b).count("1")


class _UnionFind:
    def __init__(self, n):
        self.parent = list(range(n))

    def find(self, x):
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[ra] = rb


def cluster_by_template(rows):
    """Return groups of 2+ rows whose body_template_fingerprint values are
    within TEMPLATE_HAMMING_THRESHOLD bits of each other."""
    fps = []
    for r in rows:
        fp_hex = r.get("body_template_fingerprint") or ""
        fps.append(int(fp_hex, 16) if fp_hex else None)

    indices = [i for i, fp in enumerate(fps) if fp]  # skip empty/all-zero bodies
    uf = _UnionFind(len(rows))
    for a in range(len(indices)):
        i = indices[a]
        for b in range(a + 1, len(indices)):
            j = indices[b]
            if _hamming(fps[i], fps[j]) <= TEMPLATE_HAMMING_THRESHOLD:
                uf.union(i, j)

    clusters = defaultdict(list)
    for i in indices:
        clusters[uf.find(i)].append(rows[i])
    return [members for members in clusters.values() if len(members) >= 2]


FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF")
CAT_FILLS = {
    "likely_scam": PatternFill("solid", fgColor="F8CBAD"),
    "review": PatternFill("solid", fgColor="FFE699"),
    "bulk_marketing": PatternFill("solid", fgColor="E2EFDA"),
}
CELL_FONT = Font(name=FONT, size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


def main():
    if not os.path.exists(MASTER_CSV):
        print(f"No {MASTER_CSV} yet -- nothing to build.")
        return

    with open(MASTER_CSV, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    url_rows = []
    if os.path.exists(MASTER_URLS_CSV):
        with open(MASTER_URLS_CSV, encoding="utf-8") as f:
            url_rows = list(csv.DictReader(f))
    enrichment = load_enrichment()

    wb = Workbook()

    # ---- Summary ----
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Scam Email Tracker — Master Report"
    ws["B2"].font = Font(name=FONT, bold=True, size=16)
    ws["B3"] = f"{len(rows)} messages total"
    ws["B3"].font = Font(name=FONT, italic=True, size=10, color="666666")

    cat_counts = Counter(r["category"] for r in rows)
    row = 5
    ws.cell(row=row, column=2, value="Category").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=3, value="Count").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    for cat in ["likely_scam", "review", "bulk_marketing"]:
        row += 1
        ws.cell(row=row, column=2, value=cat).font = CELL_FONT
        ws.cell(row=row, column=3, value=cat_counts.get(cat, 0)).font = CELL_FONT

    row += 2
    domain_counter = Counter(r["from_domain"] for r in rows if r["from_domain"])
    ws.cell(row=row, column=2, value="Repeated sender domains (2+ messages)").font = Font(name=FONT, bold=True, size=11)
    row += 1
    ws.cell(row=row, column=2, value="Domain").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=3, value="Count").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    for d, c in domain_counter.most_common(20):
        if c < 2:
            continue
        row += 1
        ws.cell(row=row, column=2, value=d).font = CELL_FONT
        ws.cell(row=row, column=3, value=c).font = CELL_FONT

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14

    # ---- Emails (all) ----
    ws2 = wb.create_sheet("Emails")
    fieldnames = list(rows[0].keys())
    for col, name in enumerate(fieldnames, start=1):
        c = ws2.cell(row=1, column=col, value=name.replace("_", " ").title())
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for r_idx, r in enumerate(rows, start=2):
        cat_fill = CAT_FILLS.get(r.get("category", ""))
        for col, name in enumerate(fieldnames, start=1):
            c = ws2.cell(row=r_idx, column=col, value=r[name])
            c.font = CELL_FONT
            c.alignment = WRAP
            if name == "category" and cat_fill:
                c.fill = cat_fill
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    widths = {"date": 20, "subject": 40, "from_name": 16, "from_addr": 28, "from_domain": 22,
              "return_path": 28, "reply_to_addr": 24, "dkim": 8, "spf": 8, "dmarc": 8,
              "originating_ips": 24, "url_count": 10, "url_domains": 26, "phone_numbers": 20,
              "btc_addresses": 20, "eth_addresses": 20, "contact_emails_in_body": 30,
              "first_url": 30, "id": 12, "source": 14, "category": 16, "list_unsubscribe": 14,
              "impersonated_brands": 24, "body_template_fingerprint": 18,
              "scam_score": 10, "scam_signals": 50, "date_iso": 12,
              "link_mismatches": 32, "url_flags": 24, "remote_image_domains": 26,
              "image_count": 10, "image_hashes": 24, "image_phashes": 24,
              "image_filenames": 24, "image_ocr_chars": 10}
    for col, name in enumerate(fieldnames, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = widths.get(name, 16)

    # ---- URLs ----
    ws3 = wb.create_sheet("URLs")
    url_fields = ["id", "url", "domain"]
    for col, name in enumerate(url_fields, start=1):
        c = ws3.cell(row=1, column=col, value=name.title())
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for r_idx, r in enumerate(url_rows, start=2):
        for col, name in enumerate(url_fields, start=1):
            ws3.cell(row=r_idx, column=col, value=r[name]).font = CELL_FONT
    ws3.freeze_panes = "A2"
    if url_rows:
        ws3.auto_filter.ref = ws3.dimensions
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 70
    ws3.column_dimensions["C"].width = 30

    # ---- Domain Clusters ----
    ws4 = wb.create_sheet("Domain Clusters")
    cluster_map = defaultdict(list)
    for r in rows:
        if r["from_domain"]:
            cluster_map[r["from_domain"]].append(r)
    headers4 = ["Sender Domain", "Message Count", "Categories Seen", "Registered", "Registrar",
                "Sample Subjects", "Sample Message IDs"]
    for col, name in enumerate(headers4, start=1):
        c = ws4.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r_idx = 2
    ENRICHED_FILL = PatternFill("solid", fgColor="D9E1F2")
    for d, items in sorted(cluster_map.items(), key=lambda x: -len(x[1])):
        if len(items) < 2:
            continue
        cats = ", ".join(sorted(set(it["category"] for it in items)))
        subs = " | ".join(it["subject"][:60] for it in items[:4])
        ids = ", ".join(it["id"] for it in items[:6])
        enr = enrichment.get(d)
        registered = enr["registered_date"] if enr else ""
        registrar = enr["registrar"] if enr else ""
        ws4.cell(row=r_idx, column=1, value=d).font = CELL_FONT
        ws4.cell(row=r_idx, column=2, value=len(items)).font = CELL_FONT
        ws4.cell(row=r_idx, column=3, value=cats).font = CELL_FONT
        c_reg = ws4.cell(row=r_idx, column=4, value=registered)
        c_reg.font = CELL_FONT
        c_regr = ws4.cell(row=r_idx, column=5, value=registrar)
        c_regr.font = CELL_FONT
        if enr:
            c_reg.fill = ENRICHED_FILL
            c_regr.fill = ENRICHED_FILL
        ws4.cell(row=r_idx, column=6, value=subs).font = CELL_FONT
        ws4.cell(row=r_idx, column=6).alignment = WRAP
        ws4.cell(row=r_idx, column=7, value=ids).font = CELL_FONT
        r_idx += 1
    ws4.freeze_panes = "A2"
    if r_idx > 2:
        ws4.auto_filter.ref = f"A1:G{r_idx-1}"
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 14
    ws4.column_dimensions["C"].width = 24
    ws4.column_dimensions["D"].width = 14
    ws4.column_dimensions["E"].width = 20
    ws4.column_dimensions["F"].width = 60
    ws4.column_dimensions["G"].width = 40

    # ---- Domain Enrichment (WHOIS notes) ----
    ws5 = wb.create_sheet("Domain Enrichment")
    headers5 = ["Domain", "Registered", "Registrar", "Nameservers",
                "Registrant Org", "Country", "Notes"]
    for col, name in enumerate(headers5, start=1):
        c = ws5.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r_idx = 2
    for d, enr in enrichment.items():
        values5 = [d, enr.get("registered_date", ""), enr.get("registrar", ""),
                   enr.get("nameservers", ""), enr.get("registrant_org", ""),
                   enr.get("registrant_country", ""), enr.get("notes", "")]
        for col, val in enumerate(values5, start=1):
            c = ws5.cell(row=r_idx, column=col, value=val)
            c.font = CELL_FONT
            if col in (4, 7):
                c.alignment = WRAP
        r_idx += 1
    ws5.freeze_panes = "A2"
    if r_idx > 2:
        ws5.auto_filter.ref = f"A1:G{r_idx-1}"
    for col_letter, w_ in zip("ABCDEFG", (26, 14, 20, 40, 22, 10, 60)):
        ws5.column_dimensions[col_letter].width = w_

    # ---- Registration Clusters (domains that were bought together) ----
    # Groups enriched domains sharing registrar + nameserver infrastructure.
    # Burner domains registered through the same registrar, parked on the
    # same nameservers, within days of each other, are one operator's
    # shopping trip -- a stronger link than anything in the mail itself.
    ws_reg = wb.create_sheet("Registration Clusters")

    def ns_roots(ns_field):
        roots = set()
        for ns in (ns_field or "").split(";"):
            ns = ns.strip().lower()
            if ns:
                parts = ns.split(".")
                if len(parts) >= 2:
                    roots.add(".".join(parts[-2:]))
        return frozenset(roots)

    reg_clusters = defaultdict(list)
    for d, enr in enrichment.items():
        registrar = (enr.get("registrar") or "").strip().lower()
        roots = ns_roots(enr.get("nameservers"))
        if registrar and roots:
            reg_clusters[(registrar, roots)].append((d, enr))
    headers_reg = ["Registrar", "Nameserver Infrastructure", "Domains", "Domain Count",
                   "Registration Dates", "Span (days)", "Messages From These Domains"]
    for col, name in enumerate(headers_reg, start=1):
        c = ws_reg.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    msg_count_by_domain = Counter(r["from_domain"] for r in rows if r["from_domain"])
    r_idx = 2
    for (registrar, roots), members in sorted(reg_clusters.items(), key=lambda x: -len(x[1])):
        if len(members) < 2:
            continue
        doms = sorted(d for d, _ in members)
        dates = sorted(e.get("registered_date", "") for _, e in members if e.get("registered_date"))
        span = ""
        if len(dates) >= 2:
            try:
                span = (date.fromisoformat(dates[-1][:10])
                        - date.fromisoformat(dates[0][:10])).days
            except ValueError:
                pass
        total_msgs = sum(msg_count_by_domain.get(d, 0) for d in doms)
        values_reg = [registrar, "; ".join(sorted(roots)), "; ".join(doms), len(doms),
                      ", ".join(dates), span, total_msgs]
        for col, val in enumerate(values_reg, start=1):
            c = ws_reg.cell(row=r_idx, column=col, value=val)
            c.font = CELL_FONT
            if col in (2, 3, 5):
                c.alignment = WRAP
        r_idx += 1
    ws_reg.freeze_panes = "A2"
    if r_idx > 2:
        ws_reg.auto_filter.ref = f"A1:G{r_idx-1}"
    for col_letter, w_ in zip("ABCDEFG", (24, 30, 50, 12, 26, 10, 14)):
        ws_reg.column_dimensions[col_letter].width = w_

    # ---- Origin ASNs (where the mail actually comes from) ----
    # Each message's best-guess origin IP (deepest public hop in its
    # Received chain) mapped to its ASN/hosting provider via
    # data/ip_enrichment.csv (lookup_ips.py). Concentration here is a
    # reportable pattern: a hoster carrying a big share of your scam volume
    # has an abuse desk that wants (or needs) to know. Note the caveat:
    # the Received chain can be forged below the hop your own provider
    # recorded, so treat origins as strong hints, not proof.
    ip_enrichment = load_ip_enrichment()
    ws_asn = wb.create_sheet("Origin ASNs")
    asn_map = defaultdict(list)   # asn -> [(row, ip)]
    unenriched_ips = set()
    for r in rows:
        ip = origin_ip_of(r)
        if not ip:
            continue
        enr = ip_enrichment.get(ip)
        if enr and enr.get("asn"):
            asn_map[enr["asn"]].append((r, ip))
        else:
            unenriched_ips.add(ip)
    headers_a = ["ASN", "AS Name", "ISP", "Countries", "Messages", "Likely Scam",
                 "Distinct IPs", "Distinct Sender Domains", "Sample Subjects"]
    for col, name in enumerate(headers_a, start=1):
        c = ws_asn.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r_idx = 2
    for asn, entries in sorted(asn_map.items(), key=lambda x: -len(x[1])):
        items = [r for r, _ in entries]
        ips = sorted(set(ip for _, ip in entries))
        enr0 = ip_enrichment.get(ips[0], {})
        countries = sorted(set(
            ip_enrichment.get(ip, {}).get("country_code", "") for ip in ips
        ) - {""})
        doms = sorted(set(it["from_domain"] for it in items if it["from_domain"]))
        n_scam = sum(1 for it in items if it["category"] == "likely_scam")
        subs = " | ".join(it["subject"][:60] for it in items[:4])
        values_a = [asn, enr0.get("as_name", ""), enr0.get("isp", ""),
                    ", ".join(countries), len(items), n_scam, len(ips), len(doms), subs]
        for col, val in enumerate(values_a, start=1):
            c = ws_asn.cell(row=r_idx, column=col, value=val)
            c.font = CELL_FONT
            if col == 9:
                c.alignment = WRAP
        r_idx += 1
    ws_asn.freeze_panes = "A2"
    if r_idx > 2:
        ws_asn.auto_filter.ref = f"A1:I{r_idx-1}"
    for col_letter, w_ in zip("ABCDEFGHI", (10, 24, 24, 12, 11, 11, 12, 20, 60)):
        ws_asn.column_dimensions[col_letter].width = w_
    if unenriched_ips:
        c = ws_asn.cell(row=r_idx + 1, column=1,
                        value=f"{len(unenriched_ips)} origin IP(s) not yet enriched -- "
                              "run lookup_ips.py and rebuild.")
        c.font = Font(name=FONT, italic=True, size=10, color="666666")

    # ---- Template Clusters (messages sharing a body template) ----
    # Groups messages whose bodies are near-duplicates once the parts that
    # vary per-blast (URLs, emails, phones, wallets) are stripped out -- this
    # catches the same scam template being reused across different
    # senders/domains even when there's no exact repeated contact string.
    ws_templates = wb.create_sheet("Template Clusters")
    headers_t = ["Cluster", "Message Count", "Distinct Sender Domains", "Categories Seen",
                 "Contacts Seen In Cluster", "Sample Subjects", "Sample Message IDs"]
    for col, name in enumerate(headers_t, start=1):
        c = ws_templates.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    template_clusters = sorted(cluster_by_template(rows), key=lambda members: -len(members))
    r_idx = 2
    for cluster_num, members in enumerate(template_clusters, start=1):
        doms = sorted(set(m["from_domain"] for m in members if m["from_domain"]))
        cats = ", ".join(sorted(set(m["category"] for m in members)))
        contacts = set()
        for m in members:
            contacts.update(x.strip() for x in (m.get("contacts_functional") or "").split(";") if x.strip())
            contacts.update(x.strip() for x in (m.get("phones_functional") or "").split(";") if x.strip())
        subs = " | ".join(m["subject"][:60] for m in members[:4])
        ids = ", ".join(m["id"] for m in members[:6])
        ws_templates.cell(row=r_idx, column=1, value=cluster_num).font = CELL_FONT
        ws_templates.cell(row=r_idx, column=2, value=len(members)).font = CELL_FONT
        ws_templates.cell(row=r_idx, column=3, value=len(doms)).font = CELL_FONT
        ws_templates.cell(row=r_idx, column=4, value=cats).font = CELL_FONT
        c_contacts = ws_templates.cell(row=r_idx, column=5, value="; ".join(sorted(contacts)))
        c_contacts.font = CELL_FONT
        c_contacts.alignment = WRAP
        c_subs = ws_templates.cell(row=r_idx, column=6, value=subs)
        c_subs.font = CELL_FONT
        c_subs.alignment = WRAP
        ws_templates.cell(row=r_idx, column=7, value=ids).font = CELL_FONT
        r_idx += 1
    ws_templates.freeze_panes = "A2"
    if r_idx > 2:
        ws_templates.auto_filter.ref = f"A1:G{r_idx-1}"
    ws_templates.column_dimensions["A"].width = 10
    ws_templates.column_dimensions["B"].width = 14
    ws_templates.column_dimensions["C"].width = 20
    ws_templates.column_dimensions["D"].width = 24
    ws_templates.column_dimensions["E"].width = 40
    ws_templates.column_dimensions["F"].width = 60
    ws_templates.column_dimensions["G"].width = 40

    # ---- Impersonated Brands (subcategory on top of triage) ----
    ws_brands = wb.create_sheet("Brands")
    brand_map = defaultdict(list)
    for r in rows:
        for b in (r.get("impersonated_brands") or "").split(";"):
            b = b.strip()
            if b:
                brand_map[b].append(r)
    headers_b = ["Brand", "Message Count", "Categories Seen", "Distinct Sender Domains",
                 "Sample Subjects", "Sample Message IDs"]
    for col, name in enumerate(headers_b, start=1):
        c = ws_brands.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r_idx = 2
    for b, items in sorted(brand_map.items(), key=lambda x: -len(x[1])):
        doms = sorted(set(it["from_domain"] for it in items if it["from_domain"]))
        cats = ", ".join(sorted(set(it["category"] for it in items)))
        subs = " | ".join(it["subject"][:60] for it in items[:4])
        ids = ", ".join(it["id"] for it in items[:6])
        ws_brands.cell(row=r_idx, column=1, value=b).font = CELL_FONT
        ws_brands.cell(row=r_idx, column=2, value=len(items)).font = CELL_FONT
        ws_brands.cell(row=r_idx, column=3, value=cats).font = CELL_FONT
        ws_brands.cell(row=r_idx, column=4, value=len(doms)).font = CELL_FONT
        c_subs = ws_brands.cell(row=r_idx, column=5, value=subs)
        c_subs.font = CELL_FONT
        c_subs.alignment = WRAP
        ws_brands.cell(row=r_idx, column=6, value=ids).font = CELL_FONT
        r_idx += 1
    ws_brands.freeze_panes = "A2"
    if r_idx > 2:
        ws_brands.auto_filter.ref = f"A1:F{r_idx-1}"
    ws_brands.column_dimensions["A"].width = 26
    ws_brands.column_dimensions["B"].width = 14
    ws_brands.column_dimensions["C"].width = 24
    ws_brands.column_dimensions["D"].width = 20
    ws_brands.column_dimensions["E"].width = 60
    ws_brands.column_dimensions["F"].width = 40

    # ---- Phone Numbers ----
    # Confidence: "functional" = keyword-confirmed (near "call"/"phone"/etc.)
    # AND not sitting in a dense cluster of other PII -- these are worth
    # acting on. "filler" = matched by formatting alone or found inside a
    # stuffed word-salad block -- probably spam-filter evasion noise, not a
    # real contact point. See ioc_lib.classify_phones.
    FUNCTIONAL_FILL = PatternFill("solid", fgColor="C6EFCE")
    FILLER_FILL = PatternFill("solid", fgColor="F2F2F2")
    FUNCTIONAL_FONT = Font(name=FONT, size=10, bold=True, color="006100")
    FILLER_FONT = Font(name=FONT, size=10, italic=True, color="808080")

    ws7 = wb.create_sheet("Phone Numbers")
    phone_map = defaultdict(list)
    for r in rows:
        for p in (r.get("phone_numbers") or "").split(";"):
            p = p.strip()
            if p:
                functional_set = {x.strip() for x in (r.get("phones_functional") or "").split(";") if x.strip()}
                phone_map[p].append((r, p in functional_set))
    headers7 = ["Phone Number", "Confidence", "Times Seen", "Distinct Sender Domains",
                "Sample Subjects", "Sample Message IDs"]
    for col, name in enumerate(headers7, start=1):
        c = ws7.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r_idx = 2
    # sort: functional first, then by frequency
    def phone_sort_key(item):
        p, entries = item
        is_func = any(f for _, f in entries)
        return (0 if is_func else 1, -len(entries))
    for p, entries in sorted(phone_map.items(), key=phone_sort_key):
        items = [r for r, f in entries]
        is_functional = any(f for _, f in entries)
        doms = sorted(set(it["from_domain"] for it in items if it["from_domain"]))
        subs = " | ".join(it["subject"][:60] for it in items[:4])
        ids = ", ".join(it["id"] for it in items[:6])
        conf_label = "functional" if is_functional else "filler"
        conf_fill = FUNCTIONAL_FILL if is_functional else FILLER_FILL
        conf_font = FUNCTIONAL_FONT if is_functional else FILLER_FONT
        ws7.cell(row=r_idx, column=1, value=p).font = CELL_FONT
        c_conf = ws7.cell(row=r_idx, column=2, value=conf_label)
        c_conf.font = conf_font
        c_conf.fill = conf_fill
        ws7.cell(row=r_idx, column=3, value=len(items)).font = CELL_FONT
        ws7.cell(row=r_idx, column=4, value=len(doms)).font = CELL_FONT
        ws7.cell(row=r_idx, column=5, value=subs).font = CELL_FONT
        ws7.cell(row=r_idx, column=5).alignment = WRAP
        ws7.cell(row=r_idx, column=6, value=ids).font = CELL_FONT
        r_idx += 1
    ws7.freeze_panes = "A2"
    if r_idx > 2:
        ws7.auto_filter.ref = f"A1:F{r_idx-1}"
    ws7.column_dimensions["A"].width = 20
    ws7.column_dimensions["B"].width = 12
    ws7.column_dimensions["C"].width = 12
    ws7.column_dimensions["D"].width = 20
    ws7.column_dimensions["E"].width = 60
    ws7.column_dimensions["F"].width = 40

    # ---- Email Addresses (all contact emails found in body text) ----
    # Every email address extracted from message bodies, regardless of how
    # many different sender domains it showed up in (Cross-Domain Contacts
    # below only lists ones seen across 2+ domains -- this tab is the full
    # list, single-occurrence addresses included, mirroring Phone Numbers.
    ws8 = wb.create_sheet("Email Addresses")
    email_map = defaultdict(list)
    for r in rows:
        functional_set = {x.strip().lower() for x in (r.get("contacts_functional") or "").split(";") if x.strip()}
        for e in (r.get("contact_emails_in_body") or "").split(";"):
            e = e.strip()
            if e:
                email_map[e].append((r, e.lower() in functional_set))
    headers8 = ["Email Address", "Confidence", "Times Seen", "Distinct Sender Domains",
                "Sample Subjects", "Sample Message IDs"]
    for col, name in enumerate(headers8, start=1):
        c = ws8.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r_idx = 2

    def email_sort_key(item):
        e, entries = item
        is_func = any(f for _, f in entries)
        return (0 if is_func else 1, -len(entries))

    for e, entries in sorted(email_map.items(), key=email_sort_key):
        items = [r for r, f in entries]
        is_functional = any(f for _, f in entries)
        doms = sorted(set(it["from_domain"] for it in items if it["from_domain"]))
        subs = " | ".join(it["subject"][:60] for it in items[:4])
        ids = ", ".join(it["id"] for it in items[:6])
        conf_label = "functional" if is_functional else "filler"
        conf_fill = FUNCTIONAL_FILL if is_functional else FILLER_FILL
        conf_font = FUNCTIONAL_FONT if is_functional else FILLER_FONT
        ws8.cell(row=r_idx, column=1, value=e).font = CELL_FONT
        c_conf = ws8.cell(row=r_idx, column=2, value=conf_label)
        c_conf.font = conf_font
        c_conf.fill = conf_fill
        ws8.cell(row=r_idx, column=3, value=len(items)).font = CELL_FONT
        ws8.cell(row=r_idx, column=4, value=len(doms)).font = CELL_FONT
        ws8.cell(row=r_idx, column=5, value=subs).font = CELL_FONT
        ws8.cell(row=r_idx, column=5).alignment = WRAP
        ws8.cell(row=r_idx, column=6, value=ids).font = CELL_FONT
        r_idx += 1
    ws8.freeze_panes = "A2"
    if r_idx > 2:
        ws8.auto_filter.ref = f"A1:F{r_idx-1}"
    ws8.column_dimensions["A"].width = 30
    ws8.column_dimensions["B"].width = 12
    ws8.column_dimensions["C"].width = 12
    ws8.column_dimensions["D"].width = 20
    ws8.column_dimensions["E"].width = 60
    ws8.column_dimensions["F"].width = 40

    # ---- Wallets (crypto addresses found in message bodies) ----
    # Enriched with on-chain activity from lookup_wallets.py when available.
    # A wallet that has actually received funds means the campaign has
    # victims -- highlighted red, and the single most report-worthy thing in
    # this workbook (file it with IC3 / chainabuse.com).
    ACTIVE_FILL = PatternFill("solid", fgColor="FFC7CE")
    ACTIVE_FONT = Font(name=FONT, size=10, bold=True, color="9C0006")
    wallet_enrichment = load_wallet_enrichment()
    ws_wallets = wb.create_sheet("Wallets")
    wallet_map = defaultdict(list)
    for r in rows:
        for w in (r.get("btc_addresses") or "").split(";"):
            w = w.strip()
            if w:
                wallet_map[("BTC", w)].append(r)
        for w in (r.get("eth_addresses") or "").split(";"):
            w = w.strip()
            if w:
                wallet_map[("ETH", w)].append(r)
    headers_w = ["Type", "Address", "Times Seen", "Distinct Sender Domains",
                 "On-Chain Txs", "Total Received", "Balance", "First Seen",
                 "Last Seen", "Checked", "Sample Subjects", "Sample Message IDs"]
    for col, name in enumerate(headers_w, start=1):
        c = ws_wallets.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    def wallet_sort_key(item):
        (kind, w), items = item
        enr = wallet_enrichment.get(w, {})
        try:
            txs = int(enr.get("tx_count") or 0)
        except ValueError:
            txs = 0
        return (0 if txs else 1, -txs, -len(items))  # active on-chain first

    r_idx = 2
    for (kind, w), items in sorted(wallet_map.items(), key=wallet_sort_key):
        doms = sorted(set(it["from_domain"] for it in items if it["from_domain"]))
        subs = " | ".join(it["subject"][:60] for it in items[:4])
        ids = ", ".join(it["id"] for it in items[:6])
        enr = wallet_enrichment.get(w, {})
        try:
            txs = int(enr.get("tx_count") or 0)
        except ValueError:
            txs = 0
        is_active = txs > 0
        ws_wallets.cell(row=r_idx, column=1, value=kind).font = CELL_FONT
        c_addr = ws_wallets.cell(row=r_idx, column=2, value=w)
        c_addr.font = ACTIVE_FONT if is_active else CELL_FONT
        if is_active:
            c_addr.fill = ACTIVE_FILL
        ws_wallets.cell(row=r_idx, column=3, value=len(items)).font = CELL_FONT
        ws_wallets.cell(row=r_idx, column=4, value=len(doms)).font = CELL_FONT
        enr_values = [enr.get("tx_count", "") if enr else "",
                      enr.get("total_received", ""), enr.get("balance", ""),
                      enr.get("first_seen", ""), enr.get("last_seen", ""),
                      enr.get("checked_date", "")]
        for offset, val in enumerate(enr_values):
            c_enr = ws_wallets.cell(row=r_idx, column=5 + offset, value=val)
            c_enr.font = ACTIVE_FONT if is_active else CELL_FONT
        c_subs = ws_wallets.cell(row=r_idx, column=11, value=subs)
        c_subs.font = CELL_FONT
        c_subs.alignment = WRAP
        ws_wallets.cell(row=r_idx, column=12, value=ids).font = CELL_FONT
        r_idx += 1
    ws_wallets.freeze_panes = "A2"
    if r_idx > 2:
        ws_wallets.auto_filter.ref = f"A1:L{r_idx-1}"
    for col_letter, w_ in zip("ABCDEFGHIJKL", (8, 42, 12, 12, 12, 18, 16, 13, 13, 12, 50, 40)):
        ws_wallets.column_dimensions[col_letter].width = w_

    # ---- Images (attachment/inline image hashes) ----
    # Keyed by byte hash (sha256, truncated); a perceptual hash column links
    # recompressed/resized variants of the same artwork when imagehash was
    # installed at ingest time. A repeated image across sender domains is an
    # operator fingerprint just like a reused phone number.
    ws_img = wb.create_sheet("Images")
    image_map = defaultdict(list)   # sha -> [row, ...]
    image_phash = {}                # sha -> phash (first seen)
    image_names = defaultdict(set)  # sha -> filenames
    for r in rows:
        hashes = [h.strip() for h in (r.get("image_hashes") or "").split(";") if h.strip()]
        phashes = [p.strip() for p in (r.get("image_phashes") or "").split(";") if p.strip()]
        names = [n.strip() for n in (r.get("image_filenames") or "").split(";") if n.strip()]
        for i, h in enumerate(hashes):
            image_map[h].append(r)
            if i < len(phashes) and h not in image_phash:
                image_phash[h] = phashes[i]
            if i < len(names):
                image_names[h].add(names[i])
    headers_i = ["Image Hash (sha256)", "Perceptual Hash", "Times Seen",
                 "Distinct Sender Domains", "Filenames", "Sample Subjects", "Sample Message IDs"]
    for col, name in enumerate(headers_i, start=1):
        c = ws_img.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r_idx = 2
    for h, items in sorted(image_map.items(), key=lambda x: -len(x[1])):
        doms = sorted(set(it["from_domain"] for it in items if it["from_domain"]))
        subs = " | ".join(it["subject"][:60] for it in items[:4])
        ids = ", ".join(it["id"] for it in items[:6])
        ws_img.cell(row=r_idx, column=1, value=h).font = CELL_FONT
        ws_img.cell(row=r_idx, column=2, value=image_phash.get(h, "")).font = CELL_FONT
        ws_img.cell(row=r_idx, column=3, value=len(items)).font = CELL_FONT
        ws_img.cell(row=r_idx, column=4, value=len(doms)).font = CELL_FONT
        c_names = ws_img.cell(row=r_idx, column=5, value="; ".join(sorted(image_names[h])[:4]))
        c_names.font = CELL_FONT
        c_names.alignment = WRAP
        c_subs = ws_img.cell(row=r_idx, column=6, value=subs)
        c_subs.font = CELL_FONT
        c_subs.alignment = WRAP
        ws_img.cell(row=r_idx, column=7, value=ids).font = CELL_FONT
        r_idx += 1
    ws_img.freeze_panes = "A2"
    if r_idx > 2:
        ws_img.auto_filter.ref = f"A1:G{r_idx-1}"
    for col_letter, w in zip("ABCDEFG", (22, 20, 12, 20, 30, 60, 40)):
        ws_img.column_dimensions[col_letter].width = w

    # ---- Operator Fingerprints (contacts/phones/wallets reused across domains) ----
    # A value that shows up in messages sent from 2+ different (often
    # random-string, disposable) sender domains is a much stronger signal
    # than domain repetition alone -- it suggests the same operator/kit
    # reusing one real identifier across many burner domains. Wallets carry
    # the strongest signal (real cost to replace); emails/phones keep their
    # functional/filler confidence label since spam bodies sometimes stuff
    # in unrelated PII as noise (see ioc_lib.classify_contacts/classify_phones).
    ws6 = wb.create_sheet("Operator Fingerprints")
    headers6 = ["Type", "Value", "Confidence", "Distinct Sender Domains", "Example Sender Domains"]
    for col, name in enumerate(headers6, start=1):
        c = ws6.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    fingerprints = []  # (type, value, confidence_label, domains, sort_rank)
    for p, entries in phone_map.items():
        doms = sorted(set(it["from_domain"] for it, _ in entries if it["from_domain"]))
        if len(doms) < 2:
            continue
        is_functional = any(f for _, f in entries)
        fingerprints.append(("phone", p, "functional" if is_functional else "filler", doms,
                              1 if is_functional else 2))
    for e, entries in email_map.items():
        doms = sorted(set(it["from_domain"] for it, _ in entries if it["from_domain"]))
        if len(doms) < 2:
            continue
        is_functional = any(f for _, f in entries)
        fingerprints.append(("email", e, "functional" if is_functional else "filler", doms,
                              1 if is_functional else 2))
    for (kind, w), items in wallet_map.items():
        doms = sorted(set(it["from_domain"] for it in items if it["from_domain"]))
        if len(doms) < 2:
            continue
        fingerprints.append((kind, w, "wallet", doms, 0))
    for h, items in image_map.items():
        doms = sorted(set(it["from_domain"] for it in items if it["from_domain"]))
        if len(doms) < 2:
            continue
        fingerprints.append(("image", h, "image", doms, 1))

    fingerprints.sort(key=lambda x: (x[4], -len(x[3])))

    r_idx = 2
    for kind, value, conf_label, doms, _ in fingerprints:
        is_strong = conf_label in ("functional", "wallet", "image")
        conf_fill = FUNCTIONAL_FILL if is_strong else FILLER_FILL
        conf_font = FUNCTIONAL_FONT if is_strong else FILLER_FONT
        ws6.cell(row=r_idx, column=1, value=kind).font = CELL_FONT
        ws6.cell(row=r_idx, column=2, value=value).font = CELL_FONT
        c_conf = ws6.cell(row=r_idx, column=3, value=conf_label)
        c_conf.font = conf_font
        c_conf.fill = conf_fill
        ws6.cell(row=r_idx, column=4, value=len(doms)).font = CELL_FONT
        c_ex = ws6.cell(row=r_idx, column=5, value="; ".join(doms[:8]))
        c_ex.font = CELL_FONT
        c_ex.alignment = WRAP
        r_idx += 1
    ws6.freeze_panes = "A2"
    if r_idx > 2:
        ws6.auto_filter.ref = f"A1:E{r_idx-1}"
    ws6.column_dimensions["A"].width = 10
    ws6.column_dimensions["B"].width = 36
    ws6.column_dimensions["C"].width = 14
    ws6.column_dimensions["D"].width = 20
    ws6.column_dimensions["E"].width = 70

    # ---- Trends (per-month view -- the point of the whole exercise) ----
    # Volume by category, the month's top impersonated brands, and how many
    # sender domains appeared for the first time that month (burner-domain
    # churn). Rows whose Date header couldn't be parsed land in "(unknown)".
    ws_trends = wb.create_sheet("Trends")
    by_month = defaultdict(list)
    for r in rows:
        by_month[month_of(r) or "(unknown)"].append(r)
    domain_first_month = {}
    for m in sorted(by_month):
        for r in by_month[m]:
            d = r.get("from_domain")
            if d and d not in domain_first_month:
                domain_first_month[d] = m
    headers_tr = ["Month", "Total", "Likely Scam", "Review", "Bulk Marketing",
                  "New Sender Domains", "Top Brands"]
    for col, name in enumerate(headers_tr, start=1):
        c = ws_trends.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r_idx = 2
    for m in sorted(by_month):
        items = by_month[m]
        cats = Counter(it["category"] for it in items)
        new_domains = sum(1 for d, fm in domain_first_month.items() if fm == m)
        brand_counts = Counter()
        for it in items:
            for b in (it.get("impersonated_brands") or "").split(";"):
                if b.strip():
                    brand_counts[b.strip()] += 1
        top_brands = ", ".join(f"{b} ({n})" for b, n in brand_counts.most_common(3))
        ws_trends.cell(row=r_idx, column=1, value=m).font = CELL_FONT
        ws_trends.cell(row=r_idx, column=2, value=len(items)).font = CELL_FONT
        c_scam = ws_trends.cell(row=r_idx, column=3, value=cats.get("likely_scam", 0))
        c_scam.font = CELL_FONT
        c_scam.fill = CAT_FILLS["likely_scam"]
        ws_trends.cell(row=r_idx, column=4, value=cats.get("review", 0)).font = CELL_FONT
        ws_trends.cell(row=r_idx, column=5, value=cats.get("bulk_marketing", 0)).font = CELL_FONT
        ws_trends.cell(row=r_idx, column=6, value=new_domains).font = CELL_FONT
        c_brands = ws_trends.cell(row=r_idx, column=7, value=top_brands)
        c_brands.font = CELL_FONT
        c_brands.alignment = WRAP
        r_idx += 1
    ws_trends.freeze_panes = "A2"
    for col_letter, w in zip("ABCDEFG", (12, 10, 12, 10, 16, 20, 50)):
        ws_trends.column_dimensions[col_letter].width = w

    # ---- Summary: cross-campaign signals (filled in now that the other
    # tabs' data is available) ----
    row += 2
    ws.cell(row=row, column=2, value="Cross-campaign signals").font = Font(name=FONT, bold=True, size=11)
    row += 1
    ws.cell(row=row, column=2, value="Body-template clusters (2+ messages sharing a template)").font = CELL_FONT
    ws.cell(row=row, column=3, value=len(template_clusters)).font = CELL_FONT
    row += 1
    ws.cell(row=row, column=2, value="Operator fingerprints (contact/phone/wallet reused across domains)").font = CELL_FONT
    ws.cell(row=row, column=3, value=len(fingerprints)).font = CELL_FONT

    if brand_map:
        row += 2
        ws.cell(row=row, column=2, value="Top impersonated brands").font = Font(name=FONT, bold=True, size=11)
        row += 1
        ws.cell(row=row, column=2, value="Brand").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        ws.cell(row=row, column=3, value="Count").font = HEADER_FONT
        ws.cell(row=row, column=3).fill = HEADER_FILL
        for b, items in sorted(brand_map.items(), key=lambda x: -len(x[1]))[:10]:
            row += 1
            ws.cell(row=row, column=2, value=b).font = CELL_FONT
            ws.cell(row=row, column=3, value=len(items)).font = CELL_FONT

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Saved {OUT_XLSX} ({len(rows)} total rows, {cat_counts.get('likely_scam', 0)} flagged likely_scam)")


if __name__ == "__main__":
    main()
